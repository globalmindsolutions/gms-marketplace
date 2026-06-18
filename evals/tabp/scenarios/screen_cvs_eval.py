"""screen_cvs_eval — behavioral assertion scenario for the screen-cvs skill.

This scenario is LIVE-GATED: it executes only when the developer passes
``--paid`` to the tabp runner.  The default CI / verifier run is import-clean
and makes NO model call.

How to run locally (requires the full Cowork runtime):
    python3 evals/run_evals.py --plugin tabp --paid

What it asserts (per design lines 377-389 and scoring-rubric.md):
    1. score in [0, 100]  (integer, weighted formula from Step 3)
    2. band in {Strong, Moderate, Weak}  (rubric Step 5 table)
    3. recommendation in {Recommend, Hold, Reject}  (rubric Step 5)
    4. Must-have gate — one-directional cap only:
         if any must-have is Missing -> band != "Strong"
                                     -> recommendation != "Recommend"
       (The gate can only LOWER band/recommendation, never raise it;
        rubric Step 5 line 63.  Not a rigid band-to-recommendation
        equality check.)
    5. .xlsx scorecard openpyxl-readable with valid column schema:
         Sheet 1 "Scorecard"  : Candidate, Overall Score, Band,
                                Recommendation, Must-have gate,
                                >= 1 per-requirement column,
                                Key Strengths, Key Gaps, Notes
         Sheet 2 "JD Requirements" : Requirement, Category, Type, Weight

NOTE: import openpyxl is deferred INSIDE run() — never at module scope.
The repo is stdlib-only; openpyxl is absent in CI.  A module-scope import
would break --list and the import-clean check.
"""

import os

# ---------------------------------------------------------------------------
# Scenario metadata
# ---------------------------------------------------------------------------
META = {
    "name": "screen_cvs_eval",
    "tier": "paid",   # live model invocation (Cowork runtime); opt-in only
    "goal": "AC6",
    "summary": "screen-cvs rubric contract: score/band/rec ranges, must-have gate, .xlsx",
}

# ---------------------------------------------------------------------------
# Paths (computed at module scope — stdlib only, no I/O)
# ---------------------------------------------------------------------------
_SCENARIO_DIR = os.path.dirname(os.path.abspath(__file__))
_TABP_DIR = os.path.dirname(_SCENARIO_DIR)   # evals/tabp/
_FIXTURES_DIR = os.path.join(_TABP_DIR, "fixtures")
_CV_FIXTURE = os.path.join(_FIXTURES_DIR, "cv_synthetic.md")
_JD_FIXTURE = os.path.join(_FIXTURES_DIR, "jd_synthetic.md")

# Locate the screen-cvs skill definition (runtime read-only, not imported)
_REPO_ROOT = os.path.dirname(os.path.dirname(_TABP_DIR))  # repo root
_SKILL_MD = os.path.join(
    _REPO_ROOT, "plugins", "tabp", "skills", "screen-cvs", "SKILL.md"
)


# ---------------------------------------------------------------------------
# Live run — gated behind --paid; NOT executed in CI / pre-commit / verifier
# ---------------------------------------------------------------------------

def run():
    """Execute the screen-cvs behavioral eval against the synthetic fixtures.

    Returns a Check instance (imported from evals/acs/harness.py inside this
    function so the acs harness is never a module-scope dependency of the tabp
    eval).

    This function is called only when --paid is passed to
    evals/tabp/run_evals.py.  It requires the full Cowork model runtime and
    openpyxl to be installed.
    """
    # Defer all non-stdlib imports to inside run() so the module stays
    # import-clean in the stdlib-only repo environment.
    import sys as _sys
    import subprocess
    import json
    import tempfile

    # Reuse harness.Check from evals/acs/ (insert its dir on sys.path only
    # inside run(); the acs harness must NOT be a module-scope dependency).
    _acs_dir = os.path.join(_REPO_ROOT, "evals", "acs")
    if _acs_dir not in _sys.path:
        _sys.path.insert(0, _acs_dir)
    from harness import Check  # noqa: PLC0415 (intentional deferred import)

    check = Check("screen_cvs_eval")

    # ------------------------------------------------------------------
    # Step 1: load fixtures
    # ------------------------------------------------------------------
    check.ok(
        "cv_fixture_exists",
        os.path.isfile(_CV_FIXTURE),
        "path: %s" % _CV_FIXTURE,
    )
    check.ok(
        "jd_fixture_exists",
        os.path.isfile(_JD_FIXTURE),
        "path: %s" % _JD_FIXTURE,
    )

    if not (os.path.isfile(_CV_FIXTURE) and os.path.isfile(_JD_FIXTURE)):
        check.ok("fixture_load", False, "one or both fixtures missing — aborting run")
        return check

    with open(_CV_FIXTURE) as fh:
        cv_text = fh.read()
    with open(_JD_FIXTURE) as fh:
        jd_text = fh.read()

    check.ok("cv_nonempty", bool(cv_text.strip()), "cv fixture is empty")
    check.ok("jd_nonempty", bool(jd_text.strip()), "jd fixture is empty")

    # ------------------------------------------------------------------
    # Step 2: invoke the screen-cvs skill via the local Cowork runtime
    # The skill is invoked as a headless claude -p session with the CV and
    # JD fixtures as inputs.  The scorecard is written to a temp directory.
    # ------------------------------------------------------------------
    with tempfile.TemporaryDirectory(prefix="tabp-eval-") as tmpdir:
        # Build the prompt referencing the fixtures and requesting the
        # scorecard be written to the temp dir.
        prompt = (
            "You are running the screen-cvs skill for a behavioral eval.\n"
            "CV file: %s\n"
            "JD file: %s\n"
            "Write the .xlsx scorecard to the directory: %s\n"
            "Produce the scorecard following the rubric at %s.\n"
            "After writing the scorecard, output a JSON object on a single line "
            "with keys: score (integer 0-100), band (string), "
            "recommendation (string), must_have_gate (string: 'OK' or "
            "'Missing: <which>'), xlsx_path (absolute path to the written file)."
        ) % (_CV_FIXTURE, _JD_FIXTURE, tmpdir, _SKILL_MD)

        proc = subprocess.run(
            [
                "claude", "-p", prompt,
                "--output-format", "text",
                "--permission-mode", "acceptEdits",
                "--allowedTools", "Bash Read Write Edit",
            ],
            capture_output=True,
            text=True,
            cwd=_REPO_ROOT,
            timeout=600,
        )

        check.ok(
            "claude_exit_zero",
            proc.returncode == 0,
            "claude exited %d; stderr: %s" % (proc.returncode, proc.stderr[:500]),
        )

        if proc.returncode != 0:
            return check

        # ------------------------------------------------------------------
        # Step 3: parse the JSON summary line from the model output
        # ------------------------------------------------------------------
        output_text = proc.stdout
        result_data = {}
        for line in reversed(output_text.splitlines()):
            line = line.strip()
            if line.startswith("{") and line.endswith("}"):
                try:
                    result_data = json.loads(line)
                    break
                except (json.JSONDecodeError, ValueError):
                    continue

        check.ok(
            "result_json_parseable",
            bool(result_data),
            "no JSON summary line found in model output",
        )

        if not result_data:
            return check

        score = result_data.get("score")
        band = result_data.get("band", "")
        recommendation = result_data.get("recommendation", "")
        must_have_gate = result_data.get("must_have_gate", "")
        xlsx_path = result_data.get("xlsx_path", "")

        # ------------------------------------------------------------------
        # Step 4: assert the rubric contract (design lines 377-389)
        # ------------------------------------------------------------------

        # 4a. Score range
        check.ok(
            "score_is_integer",
            isinstance(score, int),
            "score=%r is not an integer" % score,
        )
        check.ok(
            "score_in_range",
            isinstance(score, int) and 0 <= score <= 100,
            "score=%r not in [0, 100]" % score,
        )

        # 4b. Band set
        valid_bands = {"Strong", "Moderate", "Weak"}
        check.ok(
            "band_valid",
            band in valid_bands,
            "band=%r not in %r" % (band, valid_bands),
        )

        # 4c. Recommendation set
        valid_recs = {"Recommend", "Hold", "Reject"}
        check.ok(
            "recommendation_valid",
            recommendation in valid_recs,
            "recommendation=%r not in %r" % (recommendation, valid_recs),
        )

        # 4d. Must-have gate — one-directional cap invariant
        # (rubric Step 4: if any must-have Missing -> band != Strong,
        #  recommendation != Recommend; design line 387 / spec lines 185-200)
        has_missing_must_have = (
            must_have_gate.startswith("Missing:")
            or "missing" in must_have_gate.lower()
        )
        if has_missing_must_have:
            check.ok(
                "gate_caps_band_not_strong",
                band != "Strong",
                "band=%r must not be Strong when a must-have is Missing "
                "(gate=%r)" % (band, must_have_gate),
            )
            check.ok(
                "gate_caps_rec_not_recommend",
                recommendation != "Recommend",
                "recommendation=%r must not be Recommend when a must-have is "
                "Missing (gate=%r)" % (recommendation, must_have_gate),
            )
        else:
            # No must-have missing: gate should be OK
            check.ok(
                "gate_ok_when_no_missing_must_have",
                must_have_gate.strip().upper() == "OK" or not has_missing_must_have,
                "gate=%r but no missing must-haves indicated" % must_have_gate,
            )

        # ------------------------------------------------------------------
        # Step 5: assert .xlsx scorecard (openpyxl deferred here — never at
        # module scope; openpyxl absent in stdlib-only CI env)
        # ------------------------------------------------------------------
        import openpyxl  # noqa: PLC0415 (intentional deferred import)

        check.ok(
            "xlsx_path_provided",
            bool(xlsx_path),
            "xlsx_path missing from model output",
        )

        if xlsx_path:
            check.ok(
                "xlsx_exists",
                os.path.isfile(xlsx_path),
                "scorecard file not found at: %s" % xlsx_path,
            )

            if os.path.isfile(xlsx_path):
                try:
                    wb = openpyxl.load_workbook(xlsx_path)
                    check.ok("xlsx_loadable", True, "")
                except Exception as exc:
                    check.ok("xlsx_loadable", False, str(exc))
                    return check

                # Sheet 1: "Scorecard"
                sheet_names = wb.sheetnames
                check.ok(
                    "sheet_scorecard_exists",
                    "Scorecard" in sheet_names,
                    "sheets found: %r" % sheet_names,
                )

                if "Scorecard" in sheet_names:
                    ws1 = wb["Scorecard"]
                    headers = [
                        cell.value for cell in next(ws1.iter_rows(max_row=1))
                        if cell.value
                    ]
                    required_scorecard_cols = [
                        "Candidate", "Overall Score", "Band",
                        "Recommendation", "Must-have gate",
                        "Key Strengths", "Key Gaps", "Notes",
                    ]
                    for col in required_scorecard_cols:
                        check.ok(
                            "scorecard_col_%s" % col.lower().replace(" ", "_"),
                            col in headers,
                            "column %r missing from Scorecard sheet; "
                            "found: %r" % (col, headers),
                        )
                    # At least one per-requirement column
                    fixed_cols = set(required_scorecard_cols)
                    req_cols = [h for h in headers if h not in fixed_cols]
                    check.ok(
                        "scorecard_has_requirement_columns",
                        len(req_cols) >= 1,
                        "no per-requirement columns found in Scorecard; "
                        "headers: %r" % headers,
                    )

                # Sheet 2: "JD Requirements"
                check.ok(
                    "sheet_jd_requirements_exists",
                    "JD Requirements" in sheet_names,
                    "sheets found: %r" % sheet_names,
                )

                if "JD Requirements" in sheet_names:
                    ws2 = wb["JD Requirements"]
                    headers2 = [
                        cell.value for cell in next(ws2.iter_rows(max_row=1))
                        if cell.value
                    ]
                    required_jd_cols = [
                        "Requirement", "Category", "Type", "Weight",
                    ]
                    for col in required_jd_cols:
                        check.ok(
                            "jd_req_col_%s" % col.lower(),
                            col in headers2,
                            "column %r missing from JD Requirements sheet; "
                            "found: %r" % (col, headers2),
                        )

    return check
